import sys
import cv2
import os
import glob
# import matplotlib.pyplot as plt
import numpy as np
import mysql.connector
import time
import datetime as dt
import threading
import tkinter as tk

from PyQt5.QtGui import QFont, QImage, QPixmap
from PyQt5.QtWidgets import QApplication, QWidget, QMessageBox, QMainWindow, QCheckBox, QTableWidgetItem, QPushButton
from login import Ui_Form           # loin
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtMultimedia import QCamera
from PyQt5 import QtWidgets
from PyQt5.QtCore import pyqtSignal, QObject, QTimer
from PyQt5.QtWidgets import QFrame, QLineEdit, QButtonGroup
# from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from gui_window import Ui_Gui_window              # gui_window             
from PyQt5 import QtCore
from datetime import datetime
from djitellopy import Tello
from pyzbar.pyzbar import decode

from product import Ui_Form_2                       # product_info
from connect_database import ConnectDatabase
from openpyxl import Workbook
from PIL import Image, ImageTk

db = mysql.connector.connect(user='root', password='123456789', host='127.0.0.1', database='data_login')
mycursor = db.cursor()

class LoginWindow(QWidget):     # Form login
    def __init__(self):
        super().__init__()
        self.uic1 = Ui_Form()  # Tạo đối tượng UI
        self.uic1.setupUi(self)  # Thiết lập giao diện

        self.setWindowFlag(Qt.FramelessWindowHint)  # Thiết lập cửa sổ không có viền
        self.setAttribute(Qt.WA_TranslucentBackground)  # Thiết lập nền của cửa sổ trong suốt

        # Kết nối nút nhấn Login, Register
        self.uic1.login_button.clicked.connect(self.login)
        self.uic1.toolButton_sign_up.clicked.connect(self.register)
        self.uic1.toolButton_account.clicked.connect(lambda: self.uic1.Background.setCurrentIndex(0))  # Nút Click Here Sign Up được ấn chuyển qua cửa sổ 2( Register)
        self.uic1.toolButton_clik_for_sign.clicked.connect(lambda: self.uic1.Background.setCurrentIndex(1))  # Nút Already Have a Account được ấn chuyển qua cửa sổ 1(Login)

        # Kết nối nút nhấn
        self.uic1.pushButton_close.clicked.connect(self.close)
        self.uic1.pushButton_close_1.clicked.connect(self.close)
        self.uic1.checkBox_show_pass.setChecked(False)  # Khởi tạo btn show_pass ban đầu = false
        self.uic1.checkBox_show_pass.stateChanged.connect(self.show_pass)  # Kiểm tra trạng thái thay đổi của btn show_pass

        lst = ['Admin', 'User']
        for x in lst:
            self.uic1.comboBox_admin.addItem(x)
        self.uic1.line_user.setFocus()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Down:
            self.uic1.line_password.setFocus()
        if event.key() == Qt.Key_Up:
            self.uic1.line_user.setFocus()
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:  # Khi nhấn nút enter
            self.login()

    def show_pass(self):
        if self.uic1.checkBox_show_pass.isChecked():
            self.uic1.line_password.setEchoMode(QLineEdit.Normal)  # Hiện password
        else:
            self.uic1.line_password.setEchoMode(QLineEdit.Password)  # Không hiện mật khẩu

    def login(self):
        try:
            self.uic1.msg = QtWidgets.QMessageBox()
            self.uic1.msg.setWindowTitle('Warning')

            username = self.uic1.line_user.text()
            password = self.uic1.line_password.text()

            print("User", username)
            print("Password", password)

            # Tạo table login trong data_login
            query = "SELECT * FROM login WHERE BINARY Username=%s AND Password=%s"
            mycursor.execute(query, (username, password))
            result = mycursor.fetchone()
            if result:
                self.close()
                global check_in
                if result[5] == "Admin":
                    check_in = "Admin"
                else:
                    check_in = "User"

                # datetime object containing current date and time
                now = datetime.now()
                # Tách ngày và giờ
                date = now.date()
                time = now.time()
                print(type(date))
                print("date and time =", date)
                query2 = "INSERT INTO history_access (Date, Time, Username, Role) VALUES (%s, %s, %s, %s)"
                mycursor.execute(query2, (date, time, result[2], result[5]))
                db.commit()

                self.main_win1 = MainWindow()
                self.main_win1.show()
            else:
                QMessageBox.information(self, "Message", "Incorrect Username or Password")
        except Exception as e:
            print("Error during login:", e)

    def register(self):
        try:
            fullname = self.uic1.line_fullname.text()
            username = self.uic1.line_user_1.text()
            email = self.uic1.line_email.text()
            password = self.uic1.line_password_1.text()
            role = self.uic1.comboBox_admin.currentText()

            query = "SELECT * FROM login WHERE Username=%s AND Password=%s AND Email=%s AND Fullname=%s AND Role=%s"
            mycursor.execute(query, (username, password, email, fullname, role))
            result = mycursor.fetchone()

            if result:
                self.uic1.msg = QtWidgets.QMessageBox()
                self.uic1.msg.setWindowTitle('Warning')
                self.uic1.msg.setText("Account Already Exists")
                self.uic1.msg.exec_()
            else:
                query1 = "INSERT INTO login (Fullname, Username, Email, Password, Role) VALUES (%s, %s, %s, %s, %s)"
                mycursor.execute(query1, (fullname, username, email, password, role))
                self.uic1.line_fullname.clear()
                self.uic1.line_user_1.clear()
                self.uic1.line_email.clear()
                self.uic1.line_password_1.clear()
                db.commit()
                self.uic1.Background.setCurrentIndex(0)
                self.uic1.msg = QtWidgets.QMessageBox()
                self.uic1.msg.setWindowTitle('')
                self.uic1.msg.setText("You have successfully registered")
                self.uic1.msg.exec_()
        except Exception as e:
            print("Error during registration:", e)


class MainWindow(QMainWindow):  # Form main
    def __init__(self):
        super().__init__()
        self.uic = Ui_Gui_window()  # Tao doi tuong UI
        self.uic.setupUi(self)  # Thiet lap giao dien

        # self.thread = {}  
        self.init_button() # Khai bao nut nhan
        self.checkin_calendar()
        
        # Khoi tao DroneController
        self.drone_controller = DroneController()
        self.drone_controller.update_frame.connect(self.update_image)
        self.drone_controller.message_signal.connect(self.update_message)
 
        # Khoi tao keyboard Controller
        self.keyboard_controller = None
        
    def init_button(self):
        # Khai bao nut nhan
        self.uic.pushButton_start.clicked.connect(self.take_off)
        self.uic.pushButton_stop.clicked.connect(self.land)
        self.uic.pushButton_home.clicked.connect(self.return_home)
        self.uic.pushButton_connect_drone.clicked.connect(self.connect_to_drone)
        self.uic.pushButton_run.clicked.connect(self.run_drone)

        # self.uic.pushButton_log_out.clicked.connect(self.log_out)
        self.uic.pushButton_search.clicked.connect(self.search)
        self.uic.pushButton_clear.clicked.connect(self.clear)
        self.uic.pushButton_clear_message.clicked.connect(self.clear_message)

        self.uic.pushButton_calendar_1.clicked.connect(self.calendar1)
        self.uic.pushButton_calendar_2.clicked.connect(self.calendar2)

        self.uic.pushButton_load.clicked.connect(self.load_data)
        self.uic.pushButton_report.clicked.connect(self.report)

    def clear_message(self):
        self.uic.textEdit_location.setText("")
        self.uic.textEdit_message.setText("")

    def checkin_calendar(self):
        global check_in
        if(check_in=="Admin"):
            self.uic.history.setEnabled(True)
            self.uic.main.setEnabled(True)
        else:
            self.uic.history.setEnabled(False)

        # Khai bao Calendar 1
        self.lich1 = QtWidgets.QCalendarWidget(self)  # khoi tao lich 1
        self.lich1.setWindowTitle("Calendar")
        self.lich1.move(1100, 300)
        self.lich1.resize(400,300)
        self.lich1.clicked[QtCore.QDate].connect(self.get_data1)
        self.lich1.hide()

        # Khai bao Calendar 2
        self.lich2 = QtWidgets.QCalendarWidget(self)  # Khoi tao lich 2 
        self.lich2.setWindowTitle("Calendar")
        self.lich2.move(1100, 400)
        self.lich2.resize(400,300)
        self.lich2.clicked[QtCore.QDate].connect(self.get_data2)
        self.lich2.hide()

    def get_data1(self, date):
        self.uic.dateEdit_1.setDate(date)       # Lay gia tri lich 1
        self.lich1.hide()

    def get_data2(self, date):
        self.uic.dateEdit_2.setDate(date)       # Lay gia tri lich 2
        self.lich2.hide()

    def update_message(self,message):
        self.uic.textEdit_message.append(message)

    def update_image(self, cv_img):
        """Cap nhat label voi hinh anh moi tu opencv"""
        # cv_img = cv2.flip(cv_img, 1)  # Lat hinh lai
        qt_img = self.convert_cv_qt(cv_img)
        self.uic.label_camera.setPixmap(qt_img)

    def convert_cv_qt(self, cv_img):
        rgb_image = cv2.cvtColor(cv_img,cv2.COLOR_BGR2RGB)
        rgb_image = cv2.cvtColor(rgb_image,cv2.COLOR_RGB2BGR)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w
        convert_to_Qt_format = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(800, 600, QtCore.Qt.KeepAspectRatio)
        return QPixmap.fromImage(p)
    
    def connect_to_drone(self):
        if not self.drone_controller.is_connected():
            self.drone_controller.start()
            position_text = self.uic.textEdit_location.toPlainText()
            self.drone_controller.take_location(position_text) 
        else:
            print("Drone da ket noi")

    def take_off(self):
        self.drone_controller.take_off()
        # self.uic.pushButton_start.setEnabled(False)
        # self.uic.pushButton_stop.setEnabled(True)
    
    def land(self):
        self.drone_controller.land()
        # self.uic.pushButton_start.setEnabled(True)
        # self.uic.pushButton_stop.setEnabled(False)

    def return_home(self):
        # Quay ve vi tri khoi tao
        self.drone_controller.return_home()   

    def run_drone(self):
        if self.uic.comboBox_choose_location.currentIndex() == 1:          # Chon mode Location
            position_text = self.uic.textEdit_location.toPlainText()
            self.drone_controller.take_location(position_text)          # luu position_text vao location o class drone_controller   
            if position_text.count(".") != 2:
                message = "Loi: Vui long nhap 3 gia tri theo dinh dang"
                self.update_message(message)
                return
            try:
                x_vel, y_vel, z_vel = map(int,position_text.split("."))
                self.drone_controller.go_to_position(x_vel,y_vel,z_vel)
                message = ("Drone da bay toi vi tri:",position_text)
                self.update_message(message)
                
            except:
                message = ("Loi trong qua trinh toi vi tri")
                self.update_message(message)

        elif self.uic.comboBox_choose_location.currentIndex() == 3:       # Chon mode Manual
            if self.keyboard_controller is None:
                self.keyboard_controller = TK_keyBoardThread(self.drone_controller)
                self.keyboard_controller.start()
            else:
                self.keyboard_controller.start_control = True  # dam bao bat dau dieu khien neu da tao

    def closeEvent(self, event):
        if self.keyboard_controller is not None:
            self.keyboard_controller.stop()
        self.drone_controller.stop()
        super().closeEvent(event)

    def search(self):
        if self.uic.comboBox_chose_the_mode.currentIndex() == 1:          # Chọn mode xem Account
            start = self.uic.dateEdit_1.text()
            date_str = start
            # Định dạng chuỗi ban đầu
            date_format = "%d/%m/%Y"
            # Chuyển đổi thành đối tượng datetime
            datetime_obj = datetime.strptime(date_str, date_format)
            # Chuyển đổi thành chuỗi mới với định dạng mong muốn
            new_date_str = datetime_obj.strftime("%Y-%m-%d")

            end = self.uic.dateEdit_2.text()
            date_str1 = end
            # Định dạng chuỗi ban đầu
            date_format1 = "%d/%m/%Y"
            # Chuyển đổi thành đối tượng datetime
            datetime_obj1 = datetime.strptime(date_str1, date_format1)
            # Chuyển đổi thành chuỗi mới với định dạng mong muốn
            new_date_str1 = datetime_obj1.strftime("%Y-%m-%d")

            sql = "SELECT * FROM history_access WHERE DATE(Date) BETWEEN %s AND %s"
            mycursor.execute(sql, (new_date_str, new_date_str1))
            # Lấy tất cả các bản ghi từ kết quả truy vấn
            myresult = mycursor.fetchall()

            # Tạo table với số hàng và cột
            self.uic.tableWidget_account.setRowCount(len(myresult))
            self.uic.tableWidget_account.setColumnCount(4)

            # Thiết lập font chữ
            font = QFont()
            font.setBold(True)
            font.setPointSize(12)  # set font size to 12

            # Thiết lập tiêu đề, độ rộng các cột, hàng
            self.uic.tableWidget_account.setHorizontalHeaderLabels(['Date', 'Time', 'Username', 'Role'])
            self.uic.tableWidget_account.horizontalHeader().setFont(font)         # Set font cho tiêu đề ngang
            self.uic.tableWidget_account.verticalHeader().setVisible(False)       # Ẩn phần tiêu đề dọc

            for i in range(4):
                self.uic.tableWidget_account.setColumnWidth(i, 190)
            self.uic.tableWidget_account.setStyleSheet("QTableWidget {border: 3px solid black;}")
            self.uic.tableWidget_account.setFont(font)

            table_row = 0
            for row in myresult:
                date_str3 = row[0].strftime("%d/%m/%Y")
                self.uic.tableWidget_account.setItem(table_row, 0, QTableWidgetItem(str(date_str3)))
                self.uic.tableWidget_account.item(table_row, 0).setTextAlignment(Qt.AlignCenter)
                self.uic.tableWidget_account.setItem(table_row, 1, QTableWidgetItem(str(row[1])))
                self.uic.tableWidget_account.item(table_row, 1).setTextAlignment(Qt.AlignCenter)
                self.uic.tableWidget_account.setItem(table_row, 2, QTableWidgetItem(str(row[2])))
                self.uic.tableWidget_account.item(table_row, 2).setTextAlignment(Qt.AlignCenter)
                self.uic.tableWidget_account.setItem(table_row, 3, QTableWidgetItem(row[3]))
                self.uic.tableWidget_account.item(table_row, 3).setTextAlignment(Qt.AlignCenter)
                table_row += 1

        elif self.uic.comboBox_chose_the_mode.currentIndex() == 2:       # Chọn mode xem Detect Box
            start = self.uic.dateEdit_1.text()
            date_str = start
            # Định dạng của chuỗi ban đầu
            date_format = "%d/%m/%Y"
            # Chuyển đổi thành đối tượng datetime
            datetime_obj = datetime.strptime(date_str, date_format)
            # Chuyển đổi thành chuỗi mới với định dạng mong muốn
            new_date_str = datetime_obj.strftime("%Y-%m-%d")

            end = self.uic.dateEdit_2.text()
            date_str1 = end
            # Định dạng của chuỗi ban đầu
            date_format1 = "%d/%m/%Y"
            # Chuyển đổi thành đối tượng datetime
            datetime_obj1 = datetime.strptime(date_str1, date_format1)
            # Chuyển đổi thành chuỗi mới với định dạng mong muốn
            new_date_str1 = datetime_obj1.strftime("%Y-%m-%d")
            # start_date = '2023-04-01'
            # end_date = '2023-04-26'

            sql = "SELECT * FROM detect_box WHERE DATE(Date) BETWEEN %s AND %s"
            mycursor.execute(sql, (new_date_str, new_date_str1))
            # Lấy tất cả các bản ghi từ kết quả truy vấn
            myresult = mycursor.fetchall()

            # Tạo table với số hàng và cột
            self.uic.tableWidget_account.setRowCount(len(myresult))
            self.uic.tableWidget_account.setColumnCount(4)

            # Thiết lập font chữ
            font = QFont()
            font.setBold(True)
            font.setPointSize(12)  # set font size to 12

            # Thiết lập tiêu đề, độ rộng các cột, hàng
            self.uic.tableWidget_account.setHorizontalHeaderLabels(['Date', 'Time', 'QRCode','Location'])
            self.uic.tableWidget_account.horizontalHeader().setFont(font)
            self.uic.tableWidget_account.verticalHeader().setVisible(False)

            for i in range(4):
                self.uic.tableWidget_account.setColumnWidth(i, 190)
            self.uic.tableWidget_account.setStyleSheet("QTableWidget {border: 3px solid black;}")
            self.uic.tableWidget_account.setFont(font)
            table_row = 0
            for row in myresult:
                date_str3 = row[0].strftime("%d/%m/%Y")
                self.uic.tableWidget_account.setItem(table_row, 0, QTableWidgetItem(str(date_str3)))
                self.uic.tableWidget_account.item(table_row, 0).setTextAlignment(Qt.AlignCenter)
                self.uic.tableWidget_account.setItem(table_row, 1, QTableWidgetItem(str(row[1])))
                self.uic.tableWidget_account.item(table_row, 1).setTextAlignment(Qt.AlignCenter)
                self.uic.tableWidget_account.setItem(table_row, 2, QTableWidgetItem(str(row[2])))
                self.uic.tableWidget_account.item(table_row, 2).setTextAlignment(Qt.AlignCenter)
                self.uic.tableWidget_account.setItem(table_row, 3, QTableWidgetItem(row[3]))
                self.uic.tableWidget_account.item(table_row, 3).setTextAlignment(Qt.AlignCenter)

                table_row += 1

            else:print("product")
    def clear(self):
        self.uic.tableWidget_account.clear()
        self.uic.tableWidget_account.setRowCount(0)
        self.uic.tableWidget_account.setColumnCount(0)

    def calendar1(self):
        if self.lich1.isHidden():
            self.lich1.show()
        else:
            self.lich1.hide()

    def calendar2(self):
        if self.lich2.isHidden():
            self.lich2.show()
        else:
            self.lich2.hide()

    def load_data(self):
        self.product_win = MainWindow_1()
        self.product_win.show()

    def report(self):
        if self.uic.comboBox_chose_the_mode.currentIndex()==0:
            try:
                mycursor.execute("SELECT * FROM qrcodes_info")
                result = mycursor.fetchall()

                # Thêm lại table_name ( tại khi in ra report thì bị mất tiêu đề ID, Name, Age)
                table_name = [i[0] for i in mycursor.description]
                print((table_name))
                print(result)

                # Khởi tạo excel
                wb = Workbook()
                ws = wb.active
                ws.title = "mysql_data"
                ws.append(table_name)  # bắt đầu ghi dữ liệu lên file excel
                for row in result:
                    ws.append(row)
                wb.save("Information of Products.xlsx")
                db.commit()

                QMessageBox.information(self,"Message"," Your report has been exported successfully")
            except:
                print("something is wrong")
            db.close()

        if self.uic.comboBox_chose_the_mode.currentIndex() == 1:
            try:
                mycursor.execute("SELECT * FROM history_access")
                result = mycursor.fetchall()

                # Thêm lại table_name ( tại khi in ra report thì bị mất tiêu đề ID, Name, Age)
                table_name = [i[0] for i in mycursor.description]
                print((table_name))
                print(result)

                # Khởi tạo excel
                wb = Workbook()
                ws = wb.active
                ws.title = "mysql_data"
                ws.append(table_name)  # bắt đầu ghi dữ liệu lên file excel
                for row in result:
                    ws.append(row)
                wb.save("History Access.xlsx")
                db.commit()

                QMessageBox.information(self, "Message", " Your report has been exported successfully")
            except:
                print("something is wrong")
            db.close()

        if self.uic.comboBox_chose_the_mode.currentIndex() == 2:
            try:
                mycursor.execute("SELECT * FROM detect_box")
                result = mycursor.fetchall()

                # Thêm lại table_name ( tại khi in ra report thì bị mất tiêu đề ID, Name, Age)
                table_name = [i[0] for i in mycursor.description]
                print((table_name))
                print(result)

                # Khởi tạo excel
                wb = Workbook()
                ws = wb.active
                ws.title = "mysql_data"
                ws.append(table_name)  # bắt đầu ghi dữ liệu lên file excel
                for row in result:
                    ws.append(row)
                wb.save("Detected Box.xlsx")
                db.commit()

                QMessageBox.information(self, "Message", " Your report has been exported successfully")
            except:
                print("something is wrong")
            db.close()


class DroneController(QThread):  # Control Drone
    update_frame = pyqtSignal(np.ndarray)
    message_signal = pyqtSignal(str)
    def __init__(self):
        super().__init__()
        self.tello = Tello()
        self.is_flying = False   # Co bao Drone bay ( takeoff or land)
        self.connected = False   # Co bao Drone connect
        self.running = False     # Co bao camera Drone chay 
        self.flag_oke = 0         # Co bao QR Code da duoc detect 
        self.location = ""
    def run(self):
        # Mo ket noi Drone va lay video stream
        if not self.connected:
            self.tello.connect()
            self.connected = True
        self.tello.streamon()
        self.running = True

        while self.running:
            # ret, cv_img = cap.read()
            frame = self.tello.get_frame_read().frame           # Doc anh Tello
            frame = cv2.flip(frame,1)
            qrcode_frame = self.detect_qrcode(frame)            # Xu ly anh detect qrcode
            cv_img = qrcode_frame
            self.update_frame.emit(cv_img)  # cho phep truyen tin hieu ra
            time.sleep(0.1)    # Nghi 100ms

    def stop(self):
        self.running = False
        if self.connected:
            self.tello.streamoff()  # Dung video stream 
            self.tello.end()        # Ngat ket noi voi drone
            self.connected = False
        self.terminate()    # Dung thread hoan toan 
        self.wait()

    def take_off(self):
        if not self.is_flying:
            self.tello.takeoff()
            self.is_flying = True
            message = ("Drone da cat canh")
            self.message_signal.emit(message)
        else:
            message = ("Drone da cat canh truoc do")
            self.message_signal.emit(message)
            # print("Drone da cat canh truoc do")
    
    def land(self):
        if self.is_flying:
            self.tello.land()
            self.is_flying = False
            message = ("Drone da ha canh")
            self.message_signal.emit(message)
        else:
            message = ("Drone da ha canh")
            self.message_signal.emit(message)
            # print("Drone da ha canh")
    
    def go_to_position(self,x_vel,y_vel,z_vel):
        try:
            if self.is_flying:
                x = x_vel*10
                y = y_vel*10
                z = z_vel*10
                yaw = 10
                # self.tello.go_xyz_speed(x_int,y_int,z_int,5)     # Bay toi vi tri x,y,z voi toc do 10m/s
                # Gui lenh bay toi vi tri bang toc do
                self.tello.send_rc_control(x,0,0,0)         # Di chuyen truc X ( trai/phai)
                time.sleep(2)
                # print("X_vel success")
                message = "X_vel success"
                self.message_signal.emit(message)

                self.tello.send_rc_control(0,y,0,0)        # Di chuyen truc Y ( truoc/sau)
                time.sleep(2)
                # print("Y_vel success")
                message = "Y_vel success"
                self.message_signal.emit(message)

                self.tello.send_rc_control(0,0,z,0)        # Bay ngang truc Z ( len/xuong)
                time.sleep(2)
                # print("Z_vel success")
                message = "Z_vel success"
                self.message_signal.emit(message)

                self.tello.send_rc_control(0,0,0,yaw)         # Khong di chuyen ( quay trai/ phai)
                message = "Yaw_vel success"
                self.message_signal.emit(message)
                # print("Yaw_vel success")
                time.sleep(2)

                self.tello.send_rc_control(0,0,0,0)            # Dung yen 
                time.sleep(0.5)
                message = ("Drone da bay toi vi tri",self.location)
                self.message_signal.emit(message)
                # print("Drone da bay toi vi tri")
        except:
            message = ("Drone khong the bay toi vi tri",self.location)
            self.message_signal.emit(message)
            # print("Drone khong the bay toi vi tri")  

    def take_location(self,position_text):    # nhan gia tri location tu position_text o class main_window
        self.location = position_text

    def return_home(self):
        # Quay ve vi tri khoi tao 
        self.tello.go_xyz_speed(0,0,0,0)

    def is_connected(self):
        return self.connected
    
    def keyboard_drone(self,x,y,z,yaw):
        self.tello.send_rc_control(x,y,z,yaw)
 
    def take_picture(self):
        if self.connected:
            frame = self.tello.get_frame_read().frame   # Lay frame hien tai
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f'Images/{timestamp}.jpg'
            cv2.imwrite(file_name,frame)
            time.sleep(0.1)
            self.message_signal.emit(f"Image saved as {file_name}")
    def detect_qrcode(self,frame):
        try:
            barcodes = decode(frame)
            if barcodes:
                for barcode in barcodes:
                    # Lay toa do cua hinh chu nhat bao quanh ma QR
                    points = barcode.polygon
                    pts = [(point.x, point.y) for point in points]
                    pts = np.array(pts, np.int32)
                    pts = pts.reshape((-1, 1, 2))

                    # Neu so diem khong bang 4, bo qua ma nay
                    if len(pts) == 4:
                        barcode_data = barcode.data.decode('utf-8')
                        barcode_type = barcode.type
                        text = f"{barcode_data}"

                        # Lay thoi gian hien tai 
                        now = datetime.now()
                        # Tach ngay va gio 
                        date = now.date()
                        current_time = now.time()

                        frame = cv2.polylines(frame, [pts], True, (0, 255, 0), 3)
                        cv2.putText(frame, text, (pts[0][0][0], pts[0][0][1] - 10), cv2.FONT_HERSHEY_SIMPLEX, 1,(255, 0, 0), 2, cv2.LINE_AA)

                        #Luu QR Code vao Database
                        self.save_qrcode_to_db(date, current_time,barcode_data)
            return frame
        except:
            return frame

    def save_qrcode_to_db(self, date, current_time, barcode_data): 
        threading.Thread(target=self.save_qrcode_to_db_oke, args=(date,current_time,barcode_data)).start()

    def save_qrcode_to_db_oke(self, date, current_time, barcode_data):
        location = self.location
        try :
        # Kiem tra QRCode da ton tai trong detect_box vao ngay hom nay hay chua
            query_check = "SELECT COUNT(*) FROM detect_box WHERE QRCode =%s AND Date =%s"
            mycursor.execute(query_check,(barcode_data,date))
            result = mycursor.fetchone()    

            if result[0] == 0:          # QRCode chua ton tai thi chen vao
                query_insert = "INSERT INTO detect_box (Date, Time, QRCode, Location) VALUES (%s, %s, %s, %s)"
                mycursor.execute(query_insert, (date, current_time, barcode_data, location))
                db.commit()

            else : #QRCode da ton tai, cap nhat thoi gian moi nhat
                query_update = "UPDATE detect_box SET Time = %s WHERE QRCode = %s AND Date = %s"
                mycursor.execute(query_update,(current_time,barcode_data,date))
                db.commit()

            if self.check_qrcode_exists(barcode_data):
               message = "QR Code dung vi tri va ton tai tren he thong"
               self.message_signal.emit(message)
            #    print("QR Code da ton tai:", barcode_data)
            else:
                self.message_signal.emit("QR Code vua quet khong ton tai tren he thong la: B_15082024_2")

        except mysql.connector.Error as err:
           print(f"Error: {err}")

    def check_qrcode_exists(self,barcode_data):
        try:
            query = "SELECT COUNT(*) FROM qrcodes_info WHERE qrcode = %s"
            mycursor.execute(query, (barcode_data,))
            result = mycursor.fetchone()
            return result[0] > 0
        except:
            pass


class TK_keyBoardThread(QThread):
    def __init__(self,drone_controller):
        super().__init__()
        self.start_control = False
        self.speed = 50
        self.drone_controller = drone_controller        # truyen doi tuong thay vi khoi tao moi 
 
    def xFunc_press(self, event):               # Khi duoc nhan nut 
        # forward / go_back / left / right
        if self.start_control:               # Kiem tra action flag and start_control khoi tao chua?
            
            # forward / back / left / right  (lr,fb,ud,yv)
            if event.char == 'w':               
                self.drone_controller.keyboard_drone(0,self.speed,0,0)
            elif event.char == 's':
                self.drone_controller.keyboard_drone(0,-(self.speed),0,0)
            elif event.char == 'a': 
                self.drone_controller.keyboard_drone(self.speed,0,0,0)       
            elif event.char == 'd':
                self.drone_controller.keyboard_drone(-(self.speed),0,0,0)
            
            # up / down / counterClockwise / Clockwise
            elif event.char == 'i':
                self.drone_controller.keyboard_drone(0,0,self.speed,0)
            elif event.char == 'k':
                self.drone_controller.keyboard_drone(0,0,-(self.speed),0)
            elif event.char == 'j':
                self.drone_controller.keyboard_drone(0,0,0,self.speed)
            elif event.char == 'l':
                self.drone_controller.keyboard_drone(0,0,0,-(self.speed))
            elif event.char == 'h':                # phim H : Drone bay ve vi tri ban dau va khoa Drone 
                self.drone_controller.return_home()
            
            elif event.char == 'p':
                self.drone_controller.land()    # ha canh
            elif event.char == 'n':
                self.drone_controller.take_off()
            elif event.char == 'z':             # z de chup hinh 
                self.drone_controller.take_picture()

    def run(self):
        # Kiem tra neu Drone da ket noi thanh cong, khong can ket noi lai
        if not self.drone_controller.is_connected():
            # Neu chua ket noi, tien hanh ket noi truoc khi vao mode 3
            self.drone_controller.tello.connect()
            self.drone_controller.tello.streamon()
            self.drone_controller.connected = True
        
        win = tk.Tk()
        win.title("KeyBoard Controller")
        win.geometry("640x360")

        # Chuyen doi sang dang hinh anh cuar Tkinter
        img = Image.open('img/img.jpg')
        tk_img = ImageTk.PhotoImage(img)            

        # Hien thi label va hinh anh tren cua so 
        label = tk.Label(win, image=tk_img, width=640, height=360)
        label.pack()     

        # Doc tu phim nhan, khi nhan thi keyboard duoc show 
        win.bind("<KeyPress>", self.xFunc_press)
        self.start_control = True   # Bat dau dieu khien keyboard
        win.mainloop()
        


    def stop(self):
        self.start_control = False
        self.terminate()    # Dung thread hoan toan 
        self.wait()   

class MainWindow_1(QWidget):               # Form products_info
    def __init__(self):
        super().__init__()
        self.uic = Ui_Form_2()  # Tạo đối tượng UI
        self.uic.setupUi(self)  # Thiết lập giao diện

        # Create a database connection object
        self.db = ConnectDatabase()

        # Connect UI elements to class variables
        self.qrcode_id = self.uic.lineEdit
        # # Restrict input to integers
        # self.qrcode_id.setValidator(QIntValidator)

        self.location = self.uic.lineEdit_2
        self.quantity = self.uic.lineEdit_3
        self.type = self.uic.lineEdit_4
        self.status = self.uic.lineEdit_5
        self.date = self.uic.lineEdit_6

        self.add_btn = self.uic.add_btn
        self.update_btn = self.uic.update_btn
        self.select_btn = self.uic.select_btn
        self.search_btn = self.uic.search_btn
        self.clear_btn = self.uic.clear_btn
        self.delete_btn = self.uic.delete_btn

        self.result_table = self.uic.tableWidget
        self.result_table.setSortingEnabled(False)
        self.buttons_list = self.uic.function_frame.findChildren(QPushButton)

        # Initialize signal-slot conections
        self.init_signal_slot()

        # Populate the initial data in the table and Type of goods/ Status dropdowns
        self.search_info()
        # self.update_type_status()

    def init_signal_slot(self):
        # Connect buttons to their respective functions
        self.add_btn.clicked.connect(self.add_info)
        self.search_btn.clicked.connect(self.search_info)
        self.clear_btn.clicked.connect(self.clear_form_info)
        self.select_btn.clicked.connect(self.select_info)
        self.update_btn.clicked.connect(self.update_info)
        self.delete_btn.clicked.connect(self.delete_info)

    def search_info(self):
        # Function to search for qrcode information and populate the table
        # self.update_type_status()
        qrcode_info = self.get_qrcode_info()

        qrcode_result = self.db.search_info(
            qrcode_id=str(qrcode_info["qrcode_id"]),
            location=qrcode_info["location"],
            quantity=qrcode_info["quantity"],
            date=qrcode_info["date"],
            type=qrcode_info["type"],
            status=qrcode_info["status"],
        )

        self.show_data(qrcode_result)

    def add_info(self):
        # Function to add qrcode information
        self.disable_buttons()

        qrcode_info = self.get_qrcode_info()

        if qrcode_info["qrcode_id"] and qrcode_info["location"]:
            check_result = self.check_qrcode_id(qrcode_id=str(qrcode_info["qrcode_id"]))    # Kiểm tra xem đã tồn tại qrcode hay chưa

            if check_result:
                QMessageBox.information(self,"Warning","Please input a new QR Code ID",QMessageBox.StandardButton.Ok)
                self.enable_buttons()
                return

            add_result = self.db.add_info(qrcode_id=str(qrcode_info["qrcode_id"]),
                                          location=qrcode_info["location"],
                                          quantity=qrcode_info["quantity"],
                                          date=qrcode_info["date"],
                                          type=qrcode_info["type"],
                                          status=qrcode_info["status"],
                                          )
            if add_result:
                QMessageBox.information(self,"Warning",f"Add fail:{add_result}, Please try again.",QMessageBox.StandardButton.Ok)

        else:
            QMessageBox.Information(self,"Warning","Please input QR Code ID and location",QMessageBox.StandardButton.Ok)

        self.search_info()
        self.enable_buttons()
    def clear_form_info(self):
        # Function to clear the form
        # self.update_type_status()
        self.qrcode_id.clear()
        self.qrcode_id.setEnabled(True)
        self.location.clear()
        self.quantity.clear()
        self.type.clear()
        self.status.clear()
        self.date.clear()

    def update_info(self):
        # Function to update qrcode information
        new_qrcode_info = self.get_qrcode_info()

        if new_qrcode_info["qrcode_id"]:
            update_result = self.db.update_info(
                qrcode_id=str(new_qrcode_info["qrcode_id"]),
                location=new_qrcode_info["location"],
                quantity=new_qrcode_info["quantity"],
                date=new_qrcode_info["date"],
                type=new_qrcode_info["type"],
                status=new_qrcode_info["status"],
            )

            if update_result:
                QMessageBox.information(self,"Warning", f"Fail to update the information:{update_result}, Please try again.",
                                        QMessageBox.StandardButton.Ok)
            else:
                self.search_info()

        else:
            QMessageBox.information(self,"Warning", "Please select one qrcode information to update.")

    def delete_info(self):
        select_row = self.result_table.currentRow()
        if select_row != -1:
            select_option = QMessageBox.warning(self, "Warning", "Are you sure to delete it?",
                                                QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel)
            if select_option == QMessageBox.StandardButton.Ok:
                qrcode_id = self.result_table.item(select_row, 0).text().strip()
                print(f"Deleting QR Code ID: {qrcode_id}")
                delete_result = self.db.delete_info(qrcode_id=qrcode_id)
                print(f"Delete Result: {delete_result}")
                if not delete_result:
                    self.search_info()
                else:
                    QMessageBox.information(self, "Warning",
                                            f"Fail to delete the information: {delete_result}, please try again",
                                            QMessageBox.StandardButton.Ok)
            else:
                self.search_info()
        else:
            QMessageBox.information(self, "Warning", "Please select one qrcode information to delete",
                                    QMessageBox.StandardButton.Ok)

    def select_info(self):
        # Function to select and populate student information in the form
        select_row = self.result_table.currentRow()
        if select_row != -1:      # Nếu có hàng được chọn
            self.qrcode_id.setEnabled(False)
            qrcode_id = self.result_table.item(select_row,0).text().strip()
            location = self.result_table.item(select_row, 1).text().strip()
            quantity = self.result_table.item(select_row, 2).text().strip()
            type = self.result_table.item(select_row, 3).text().strip()
            status = self.result_table.item(select_row, 4).text().strip()
            date = self.result_table.item(select_row, 5).text().strip()

            self.qrcode_id.setText(qrcode_id)
            self.location.setText(location)
            self.quantity.setText(quantity)
            self.type.setText(type)
            self.status.setText(status)
            self.date.setText(date)

        else:
            QMessageBox.information(self,"Warning", "Please select one qrcode information",
                                    QMessageBox.StandardButton.Ok)

    def disable_buttons(self):
        # Disable all the buttons
        for button in self.buttons_list:
            button.setDisabled(True)

    def enable_buttons(self):
        # enable all the buttons
        for button in self.buttons_list:
            button.setDisabled(False)

    def get_qrcode_info(self):
        # Function to retrieve qrcode information from the form
        qrcode_id = self.qrcode_id.text().strip()
        location = self.location.text().strip()
        quantity = self.quantity.text().strip()
        date = self.date.text().strip()
        type = self.type.text().strip()
        status = self.status.text().strip()

        qrcode_info = {
            "qrcode_id": qrcode_id,
            "location": location,
            "quantity": quantity,
            "date": date,
            "type": type,
            "status": status,
        }

        return qrcode_info

    def check_qrcode_id(self,qrcode_id):
        # Function to check if a qrcode_id already exists
        result = self.db.search_info(qrcode_id=qrcode_id)

        return result

    def show_data(self,result):    # Hiển thị dữ liệu trong bảng kết quả
        # Function to populate the table with qrcode information
        if result:
            self.result_table.setRowCount(0)
            self.result_table.setRowCount(len(result))

            for row, info in enumerate(result):
                info_list =[
                    info["qrcode"],
                    info["location"],
                    info["quantity"],
                    info["type"],
                    info["status"],
                    info["date"],
                ]

                for column, item in enumerate(info_list):
                    cell_item = QTableWidgetItem(str(item))
                    self.result_table.setItem(row,column,cell_item)
        else:
            self.result_table.setRowCount(0)
            return



if __name__ == "__main__":
    # Chay ung dung
    app = QApplication(sys.argv)
    login_win = LoginWindow()
    login_win.show()
    sys.exit(app.exec())